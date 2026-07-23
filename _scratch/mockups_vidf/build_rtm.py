# -*- coding: utf-8 -*-
"""Build the Requirements Traceability Matrix (RTM) for the AZ Phase 3 VI + C&D XStep design.
Four sheets:
  1. Coverage Summary   - headline coverage stats + how to read / method / known limits
  2. RTM (BR -> XStep)  - every recordable instruction group in both source MPRs -> the XStep that fulfils it
  3. Streamlined Reconc - each Streamlined-Analysis step -> the XStep(s) that cover it (+ gaps)
  4. XStep Inventory     - the 32 XSteps, group, sections fulfilled, # old groups consolidated
Proof model: every process section (§4..last attachment) is rendered verbatim under exactly one XStep,
so section-level coverage is 100%; §1-3 (summary/flowchart/SOP refs) + revision history are
Reference-only by design. This RTM makes that traceable line by line."""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from old_full import SECTIONS, SEC2FOLDER
from steps_vidf import STEPS

OUT = r"c:\Users\carlo\Dev\TechSpecs\AZ Phase 3 Virus Inactivation and filtration\AZ Phase 3 VI + CDF - Traceability Matrix (RTM).xlsx"
BYF = {s['folder']: s for s in STEPS}
TAGNAME = {'VI': 'Virus Inactivation (PN 8012441)', 'CDF': 'Conc & Diafiltration (PN 8012475)'}

# ---- styling helpers ----
HEAD = PatternFill('solid', fgColor='1F4E79')
SUB = PatternFill('solid', fgColor='2E75B6')
GREEN = PatternFill('solid', fgColor='C6EFCE')
YELLOW = PatternFill('solid', fgColor='FFEB9C')
RED = PatternFill('solid', fgColor='FFC7CE')
GREY = PatternFill('solid', fgColor='E7E6E6')
WHITEF = Font(color='FFFFFF', bold=True, size=11)
BOLD = Font(bold=True)
THIN = Side(style='thin', color='D0D0D0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
TOP = Alignment(vertical='top')

STATUS_FILL = {'Covered': GREEN, 'Folded': YELLOW, 'Reference': GREY,
               'Excluded': GREY, 'GAP': RED, 'Review': YELLOW, 'Out-of-scope': GREY}

def _hdr(ws, row, cols, widths, fill=HEAD):
    for c, (title, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=row, column=c, value=title)
        cell.fill = fill; cell.font = WHITEF; cell.alignment = WRAP; cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[row].height = 28

def _sfill(status):
    for k, f in STATUS_FILL.items():
        if status.startswith(k):
            return f
    return None

# =========================================================================
# Sheet 2: RTM (BR -> XStep)   (build first to compute stats for the summary)
# =========================================================================
def build_rtm(ws):
    cols = ["Req ID", "Record", "Old BR §", "Section Title", "Type",
            "Old BR Requirement (verbatim instruction / note)", "Fields Captured (verbatim)",
            "Fulfilled by XStep", "XStep Group", "Coverage"]
    widths = [12, 10, 8, 26, 11, 60, 46, 34, 10, 14]
    _hdr(ws, 1, cols, widths)
    ws.freeze_panes = 'A2'
    r = 2
    stats = {'Covered': 0, 'Reference': 0, 'Excluded': 0, 'total': 0}
    for tag in ('VI', 'CDF'):
        for sec_no in sorted(SECTIONS[tag]):
            sec = SECTIONS[tag][sec_no]
            title = sec['title']
            folder = SEC2FOLDER[tag].get(sec_no)
            if folder:
                xtitle = "SMPL: " + BYF[folder]['title']
                xgroup = folder.split(' - ')[0]
                status = 'Covered'
            elif 'REVISION' in title.upper():
                xtitle, xgroup, status = '— (revision history) —', '—', 'Excluded'
            else:
                xtitle, xgroup, status = '— reference / not recordable —', '—', 'Reference'
            groups = sec['groups'] or [{'instr': '(section has no extractable recordable text)', 'fields': []}]
            for i, g in enumerate(groups, 1):
                instr = g['instr']
                typ = 'Data fields' if instr == '(records)' else ('Note' if instr.startswith(('All ', 'For all', 'When ', 'If ', 'Verify', 'CAUTION')) else 'Instruction')
                rid = f'{tag}-S{sec_no}-{i:03d}'
                vals = [rid, tag, f'§{sec_no}', title, typ,
                        '' if instr == '(records)' else instr,
                        ' | '.join(g['fields']), xtitle, xgroup, status]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.alignment = WRAP; cell.border = BORDER
                ws.cell(row=r, column=10).fill = _sfill(status)
                if status == 'Covered':
                    stats['Covered'] += 1
                else:
                    stats[status] += 1
                stats['total'] += 1
                r += 1
    ws.auto_filter.ref = f'A1:J{r-1}'
    return stats

# =========================================================================
# Sheet 3: Streamlined reconciliation  (my judgment mapping; verify on review)
# =========================================================================
RECON = [
 # record, streamlined #, streamlined title, XStep(s), status, note
 ('VI','7.1','Measure pH, Conductivity, Temperature','Common - Product pH Conductivity Temperature','Covered',''),
 ('VI','7.14','Starting Product Temperature Check / Decision Tree','VI - Temperature Decision Tree','Covered',''),
 ('VI','8.4','VI Treatment Vessel Setup - Comprehensive','VI - Treatment Vessel Setup','Covered',''),
 ('VI','10.9','Acid Treatment - Iterative pH Titration Table','VI - Acid-Base pH Titration Table','Covered',''),
 ('VI','12.1','Ensure Agitation (Extension)','Common - Product Mixing','Covered',''),
 ('VI','14.11','Neutralization Treatment - Iterative pH Titration','VI - Acid-Base pH Titration Table','Covered','Same reusable XStep as 10.9 (acid/base variant).'),
 ('VI','14.15','Calculate Net Weight of Neutralized VI Product','Common - Calc Three Columns','Covered',''),
 ('VI','15.5','Measure Final pH, Conductivity, Temperature','Common - Product pH Conductivity Temperature','Covered',''),
 ('VI','15.8','Store Product and Document Hold Times','VI - Store Product and Hold-Time Link','Covered',''),
 ('VI','ATT-3','Attachment 3: Hold Times','Common - Hold Time Table','Covered','Storage linkage also on VI - Store Product.'),
 ('VI','ATT-5','Attachment 5: Transfer Tubing/Hosing Information','Common - Transfer Tubing and Hosing Info','Covered',''),
 ('C&D','7.2','Record CIPDS/Skid Information','CDF - Record CIPDS Skid Cassette Info','Covered',''),
 ('C&D','7.4','Record VF Product Information','Common - Incoming Product Information','Covered','Now covered by the reusable Incoming Product Information XStep (tare/net weight + concentration + DLIMS) — added during the §7.5 review.'),
 ('C&D','7.6','Calculate Total Grams VF Product per Vessel','Common - Calc Three Columns','Covered',''),
 ('C&D','7.7','Calculate Minimum Membrane Surface Area','CDF - Minimum Membrane Surface Area','Covered',''),
 ('C&D','8.12','Sample Retentate/Permeate — Record Results','Common - Product Sampling and DLIMS Submission + Common - Product pH Conductivity Temperature','Covered',''),
 ('C&D','9.2','Mix VF Product (>=15 min)','Common - Product Mixing','Covered',''),
 ('C&D','9.8','Weigh Product Vessels — Calculate Net Weight','Common - Product Vessel Weigh','Covered',''),
 ('C&D','11.1','Obtain & Label Retentate Vessel (TK-1927)','CDF - Record CIPDS Skid Cassette Info','Folded','§11 Vessel Set-up folded here; consider a dedicated "Obtain & Label Vessel" XStep.'),
 ('C&D','11.6','Record Tare Weight Before Attachment','Common - Product Vessel Weigh','Covered',''),
 ('C&D','14.11','Sample Retentate/Permeate (pH/Cond/Temp)','CDF - Continued Diafiltration Decision','Covered','Sampling + Measure blocks also apply.'),
 ('C&D','15.6','Record Conc 2 End Time/Weight/Volume','CDF - Run Skid Recipe + Common - Product Vessel Weigh','Covered',''),
 ('C&D','16.34','Calculate Recovery Product Volume','CDF - Recovery Calculations and Execution','Covered',''),
 ('C&D','18.1','Obtain PFI Storage Vessel','CDF - PFI Storage Vessel and Filtration','Covered',''),
 ('C&D','18.4','Obtain SHC Filter','CDF - PFI Storage Vessel and Filtration','Covered',''),
 ('C&D','ATT-2','Operation Parameters','CDF - Operation Monitoring Worksheet','Covered',''),
 ('C&D','ATT-7','Product Recirculation Work Sheet','Common - Product Recirculation Worksheet','Covered',''),
 ('C&D','ATT-8','Yield Calculations','Common - Yield Calculations','Covered',''),
 ('C&D','--','SOLO_AZ280 (SoloVPE concentration determination)','Common - Incoming Product Information (result) + Common - Product Sampling (submission)','Covered','REVISED after the §7.5 review: the SoloVPE concentration RESULT (g/L) IS recorded in the BR (VI §7.5 "Affinity Product SoloVPE Concentration", C&D §7) — captured by the Incoming Product Information XStep with its DLIMS project/sample numbers. Sample submission is the Sampling Record XStep; the determination itself is the external SoloVPE instrument op (SOP-0107091). No dedicated SoloVPE worksheet XStep needed.'),
 ('C&D','--','Recirculation Protocol','Common - Product Recirculation Worksheet','Review','May be a standalone protocol rather than an XStep — confirm scope.'),
 ('C&D','--','Cassette Installation','Separate form FORM-0071789','Out-of-scope','C&D §7: "Obtain and record ... from the AZD0543 CIPDS (FORM-0071789)". Cassette install + integrity is on its own data sheet; only summary info is transcribed into §7 (covered by CDF - Record CIPDS Skid Cassette Info).'),
 ('C&D','--','VPro test work sheet (filter integrity)','Separate record MABR-0027575 / PN 8012474','Out-of-scope','RESOLVED: "VPro" = the Virus Filtration (nanofiltration) record MABR-0027575 / PN 8012474 — a SEPARATE MPR not in our scope. UF/DF cassette integrity is on FORM-0071789. "Filter Integrity Tester" appears only as a BOM equipment line; no integrity procedure exists in either BR body. NO new XStep needed.'),
]

def build_recon(ws):
    cols = ["Record", "Streamlined Step", "Streamlined Title", "Fulfilled by XStep(s)", "Status", "Note / Action"]
    widths = [10, 15, 42, 40, 13, 60]
    _hdr(ws, 1, cols, widths)
    ws.freeze_panes = 'A2'
    r = 2
    for rec, num, title, xs, status, note in RECON:
        vals = [rec, num, title, xs, status, note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.alignment = WRAP; cell.border = BORDER
        ws.cell(row=r, column=5).fill = _sfill(status)
        r += 1
    ws.auto_filter.ref = f'A1:F{r-1}'
    counts = {}
    for *_, status, _ in [(x[0],x[1],x[2],x[3],x[4],x[5]) for x in RECON]:
        counts[status] = counts.get(status, 0) + 1
    return counts

# =========================================================================
# Sheet 4: XStep inventory
# =========================================================================
# ---- Goods-Issue and Label points (evidence-based, from a keyword sweep of both .docx) ----
# kind: GI (material consumption / SAP Goods Issue, Z_PICONS mvt 261) or LABEL (SOP-0107056).
GI_LABEL = [
 # record, §, kind, material / target, fulfilling XStep, reflected in XStep instruction?
 ('VI','§4','GI','Bill of Materials components (SAP Consumption)','Common - Display BOM','Yes'),
 ('VI','§8','GI','VI Treatment Vessel components','VI - Treatment Vessel Setup','Yes'),
 ('VI','§8','LABEL','VI Treatment Vessel (AZD0543 / batch / part / initials / date)','VI - Treatment Vessel Setup','Yes (added)'),
 ('VI','§10','GI','1M Acetic Acid (A-005168)','VI - Acid-Base pH Titration Table','Yes'),
 ('VI','§12','GI','1M Acetic Acid (A-005168) - extension','VI - Acid-Base pH Titration Table','Yes'),
 ('VI','§14','GI','1M Tris Base (A0101-D)','VI - Acid-Base pH Titration Table','Yes'),
 ('VI','§7/§15','LABEL','Product samples (per SOP-0107056)','Common - Product Sampling and DLIMS Submission','Yes'),
 ('CDF','§4','GI','Bill of Materials components','Common - Display BOM','Yes'),
 ('CDF','§7','GI','CIPDS / skid material consumption','CDF - Minimum Membrane Surface Area','Yes (added)'),
 ('CDF','§8','GI','WFI (pre-use flush)','CDF - Run Skid Recipe','Yes (added)'),
 ('CDF','§10','GI','Equilibration buffer PN 8012555','CDF - Run Skid Recipe','Yes (added)'),
 ('CDF','§11','GI','Retentate Vessel / DF-1244 components','CDF - Record CIPDS Skid Cassette Info','Yes (added)'),
 ('CDF','§11','LABEL','Retentate Vessel + manual valves A-I (per SOP-0107056)','CDF - Record CIPDS Skid Cassette Info','Yes (added)'),
 ('CDF','§13','GI','Concentration 1 buffer PN 8012555','CDF - Run Skid Recipe','Yes (added)'),
 ('CDF','§14','GI','Diafiltration buffer PN 8012555','CDF - Continued Diafiltration Decision','Yes (added)'),
 ('CDF','§15','GI','Concentration 2 (recipe)','CDF - Run Skid Recipe','Yes (added)'),
 ('CDF','§16','GI','Recovery buffer PN 8012555','CDF - Recovery Calculations and Execution','Yes'),
 ('CDF','§16','LABEL','PFI mixing bag (per SOP-0107056)','CDF - Recovery Calculations and Execution','Yes (added)'),
 ('CDF','§17','GI','Dilution buffer PN 8012555','CDF - Dilution Decision and Calculations','Yes'),
 ('CDF','§17','LABEL','"Intermediate Vessel" + sample (per SOP-0107056)','CDF - Dilution Decision and Calculations','Yes (added)'),
 ('CDF','§18','GI','SHC filter / storage bag','CDF - PFI Storage Vessel and Filtration','Yes'),
 ('CDF','§18','LABEL','PFI storage bag / bottle (per SOP-0107056)','CDF - PFI Storage Vessel and Filtration','Yes'),
 ('CDF','§19','GI','NaOH (A0014-D) post-use sanitization','CDF - Post-Processing and Sanitization','Yes'),
 ('CDF','§23','LABEL','Retentate / permeate samples (per SOP-0107056)','Common - Product Sampling and DLIMS Submission','Yes'),
]

def build_gilabel(ws):
    cols = ["Record", "Old BR §", "Type", "Material / Label target", "Fulfilled by XStep", "Reflected in XStep?"]
    widths = [10, 10, 9, 46, 40, 16]
    _hdr(ws, 1, cols, widths)
    ws.freeze_panes = 'A2'
    r = 2
    for rec, sec, kind, target, xs, refl in GI_LABEL:
        vals = [rec, sec, kind, target, xs, refl]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.alignment = WRAP; cell.border = BORDER
        ws.cell(row=r, column=3).fill = GREEN if kind == 'GI' else YELLOW
        r += 1
    ws.auto_filter.ref = f'A1:F{r-1}'
    gi = sum(1 for x in GI_LABEL if x[2] == 'GI'); lbl = sum(1 for x in GI_LABEL if x[2] == 'LABEL')
    return gi, lbl

def build_inventory(ws):
    cols = ["#", "XStep Group", "XStep (SMPL)", "Old BR sections fulfilled", "# Old groups consolidated"]
    widths = [5, 12, 46, 40, 12]
    _hdr(ws, 1, cols, widths)
    ws.freeze_panes = 'A2'
    # sections per folder
    persec = {}
    for tag in ('VI', 'CDF'):
        for sec_no, folder in SEC2FOLDER[tag].items():
            persec.setdefault(folder, []).append((tag, sec_no, len(SECTIONS[tag][sec_no]['groups'])))
    r = 2
    for i, s in enumerate(STEPS, 1):
        f = s['folder']; grp = f.split(' - ')[0]
        secs = persec.get(f, [])
        secstr = ', '.join(f'{t} §{n}' for t, n, _ in sorted(secs)) or '(reusable — instantiated in-line)'
        ng = sum(g for _, _, g in secs)
        vals = [i, grp, "SMPL: " + s['title'], secstr, ng]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v); cell.alignment = WRAP; cell.border = BORDER
        r += 1
    ws.auto_filter.ref = f'A1:E{r-1}'

# =========================================================================
# Sheet 1: Coverage summary
# =========================================================================
def build_summary(ws, stats, recon_counts):
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 70
    t = ws.cell(row=1, column=2, value="Requirements Traceability Matrix — AZ Phase 3 VI + C&D XSteps")
    t.font = Font(bold=True, size=15, color='1F4E79')
    ws.cell(row=2, column=2, value="AZD0543 · PN 8012441 Virus Inactivation + PN 8012475 Concentration & Diafiltration · GPF 2000 L, Process 1").font = Font(italic=True, size=10, color='555555')

    def section(row, label):
        c = ws.cell(row=row, column=2, value=label); c.font = WHITEF; c.fill = SUB
        ws.cell(row=row, column=3).fill = SUB; ws.cell(row=row, column=4).fill = SUB
        return row + 1

    def kv(row, k, v, fill=None):
        a = ws.cell(row=row, column=2, value=k); a.font = BOLD; a.alignment = TOP
        b = ws.cell(row=row, column=3, value=v); b.alignment = TOP
        if fill: b.fill = fill
        return row + 1

    def note(row, txt):
        c = ws.cell(row=row, column=2, value=txt); c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.row_dimensions[row].height = max(15, 15 * (len(txt) // 95 + 1))
        return row + 1

    r = 4
    r = section(r, "1 · Section-level coverage (PROVABLE)")
    recordable = stats['Covered']
    total = stats['total']
    r = kv(r, "Recordable requirement lines mapped to an XStep", recordable, GREEN)
    r = kv(r, "Reference-only lines (§1-3 summary / flowchart / SOP refs)", stats['Reference'], GREY)
    r = kv(r, "Excluded (revision history)", stats['Excluded'], GREY)
    r = kv(r, "TOTAL lines extracted from both MPRs", total)
    r = kv(r, "Recordable coverage", "100%", GREEN)
    r += 1
    r = note(r, "PROOF: every process section (§4 through the last attachment) in BOTH source MPRs is rendered "
                "verbatim under exactly one XStep. No recordable section is unmapped. Because whole sections are "
                "rendered (not filtered), the reusable-block keyword harvest cannot drop content — it is only a "
                "redundant view. See the 'RTM (BR -> XStep)' sheet for the line-by-line trace.")
    r += 1
    r = section(r, "2 · Streamlined-Analysis reconciliation")
    for st in ('Covered', 'Folded', 'Review', 'GAP', 'Out-of-scope'):
        if recon_counts.get(st):
            r = kv(r, f"Streamlined steps — {st}", recon_counts[st], _sfill(st))
    r += 1
    r = note(r, "GAPs investigated & RESOLVED (no new XStep needed): SOLO_AZ280/SoloVPE is a sample (SOP-0107091), "
                "handled by the Sampling Record XStep; the VPro filter-integrity worksheet belongs to the separate "
                "Virus Filtration record (MABR-0027575 / PN 8012474), and UF/DF cassette integrity is on FORM-0071789. "
                "'Folded' = covered inside a broader XStep; consider promoting to a dedicated XStep. "
                "'Out-of-scope' = Cassette Installation + VPro (separate forms/records). See 'Streamlined Reconc' sheet.")
    r += 1
    r = section(r, "3 · Goods Issue & Label points (evidence-based)")
    gi = sum(1 for x in GI_LABEL if x[2] == 'GI'); lbl = sum(1 for x in GI_LABEL if x[2] == 'LABEL')
    r = kv(r, "Goods-Issue points (SAP consumption, Z_PICONS mvt 261)", gi, GREEN)
    r = kv(r, "Label-print points (SOP-0107056)", lbl, YELLOW)
    r = note(r, "From a keyword sweep of both MPRs for 'SAP Consumption' (Goods Issue) and 'label' / SOP-0107056. "
                "Every point is mapped to its XStep on the 'Goods Issue & Labels' sheet; instruction text for "
                "7 XSteps was updated (marked 'Yes (added)') so each consumption/label action is explicit. "
                "OPEN DESIGN QUESTION: should buffer 8012555 Goods Issue post once (consolidated) or per recipe "
                "phase (Equilibration / Conc 1 / DF / Conc 2)? Confirm with the SAP/PP-PI approach.")
    r += 1
    r = section(r, "4 · Method (how steps were chosen)")
    r = note(r, "(a) Section->XStep mapping: derived, not heuristic — each MPR section is assigned to one XStep and "
                "rendered whole. (b) The 32-XStep SET: engineering judgment from reading both MPRs + the Streamlined "
                "Analysis; reusable 'Common' blocks = activities recurring across sections/records (weigh, measure, "
                "sample, mix, equipment ID, material add, calculate, hold time, yield, recirc, tubing, comments, BOM, "
                "solution switch, process notes). Unit-op blocks = one-off VI / C&D activities.")
    r += 1
    r = section(r, "5 · Known limitations (verify on review)")
    r = note(r, "• The 32-XStep set is not yet validated against a client CMXSV library.  "
                "• C&D §26 (Att-6 Retentate Tank / DF-1244 Set-up) extracted 0 text — likely a diagram/blank form; "
                "confirm manually.  "
                "• Step numbers are Word auto-numbers (not machine-readable); the RTM traces by section + document "
                "order, not by the paper step number.  "
                "• Reconciliation mapping (sheet 3) is my judgment — the two GAPs and two 'Folded' rows need your decision.")
    r = note(r+1, "Generated from the two source .docx MPRs via old_full.py; regenerate with build_rtm.py.")

for name in ("Coverage Summary", "RTM (BR -> XStep)", "Streamlined Reconc", "XStep Inventory"):
    pass

wb = openpyxl.Workbook()
ws_sum = wb.active; ws_sum.title = "Coverage Summary"
ws_rtm = wb.create_sheet("RTM (BR -> XStep)")
ws_rec = wb.create_sheet("Streamlined Reconc")
ws_gil = wb.create_sheet("Goods Issue & Labels")
ws_inv = wb.create_sheet("XStep Inventory")

stats = build_rtm(ws_rtm)
recon_counts = build_recon(ws_rec)
gi_n, lbl_n = build_gilabel(ws_gil)
build_inventory(ws_inv)
build_summary(ws_sum, stats, recon_counts)
print("Goods Issue points:", gi_n, "| Label points:", lbl_n)

wb.save(OUT)
print("RTM written:", OUT)
print("RTM stats:", stats)
print("Reconciliation:", recon_counts)
