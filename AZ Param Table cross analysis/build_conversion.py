# -*- coding: utf-8 -*-
"""Convert the OLD Phase-2 'Parameter Table Design' flat table into the NEW
PPTable_Design_v2 two-table (ZPP_PARAM + ZPP_PVAL) format. Per user (2026-07-15):
client-100 scope only; BIO_DAILY_SAMPLE reclassified 903->100 (MATNR 153); other
903 params discarded; MIN/MAX kept separate; PARAM_DESC authored as draft."""
import re, openpyxl
from openpyxl.styles import Font, PatternFill

SRC = "Phase 2 New xSteps (1).xlsx"
OUT = "Phase2 Parameters in PPTable_v2 Format.xlsx"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["Parameter Table Design"]
rows = [list(r) for r in ws.iter_rows(values_only=True)]
raw = [r for r in rows[3:] if r[2] not in (None, "")]
wb.close()

def clean(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v).strip()

RANGE = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*-\s*(-?\d+(?:\.\d+)?)\s*$")

def val_shape(t):
    m = RANGE.match(t)
    return ("R", "", m.group(1), m.group(2)) if m else ("S", t, "", "")

def data_type(t, sf, vt, tgt, lo, hi):
    if t == "C" and vt != "R":
        return "STR"
    text = tgt if vt == "S" else f"{lo}|{hi}"
    if "." in text:          # any decimal present -> float (handles decimal ranges typed 'C')
        return "FLT"
    try:
        s = int(float(sf))
    except Exception:
        s = 0
    return "FLT" if s > 0 else "INT"

def role(id2):
    u = id2.upper()
    if "_MAX" in u or "_MIN" in u or "_RANGE" in u:
        return "LIMIT"
    if any(k in u for k in ("TARGET", "SET_POINT", "SETPOINT", "TRIGGER", "WEIGHT", "NUM_FLASKS", "REQ_QTY", "_QTY")):
        return "SETPOINT"
    return "REFERENCE"

GROUP_UOP = {"FLASK_NUMBER": "INOCULUM"}
def uop(id1):
    return GROUP_UOP.get(id1, "BIOREACTOR")

GROUP_XSTEP = {
    "BIO_ANTIFOAM": "SMPL: Antifoam Addition",
    "BIO_DAILY_SAMPLE": "SMPL: Rocker Bag Daily Sample / Daily Observations",
    "FLASK_NUMBER": "SMPL: Flask Number",
}
DAILY = {
    "GF_TRIGGER": "Glucose feed trigger threshold",
    "GF_TARGET": "Glucose feed target concentration",
    "BIOREACTOR_TEMP_RANGE": "Bioreactor temperature operating range",
    "AGITATION_RANGE": "Bioreactor agitation (stir speed) operating range",
    "BIOREACTOR_PH_RANGE": "Bioreactor pH operating range",
    "PO2_RANGE": "Dissolved oxygen (pO2) operating range",
    "PRESSURE_RANGE": "Bioreactor headspace pressure operating range",
    "SPARGE_AIR_RANGE": "Sparge air flow operating range",
    "OVERLAY_AIR_RANGE": "Overlay air flow operating range",
}

def describe(id1, id2, vtxt):
    u = id2.upper()
    if id1 == "FLASK_NUMBER":
        m = re.match(r"AVERAGE_VCD_RANGE(\d+)", u)
        if m:
            return f"Seed culture average VCD bucket {m.group(1)} ({vtxt} x10^6 cells/mL) used to select flask count/volume"
        m = re.match(r"([\d.]+-[\d.]+)_WEIGHT", id2)
        if m:
            return f"Transfer weight when seed average VCD is {m.group(1)} x10^6 cells/mL"
        m = re.match(r"([\d.]+-[\d.]+)_NUM_FLASKS", id2)
        if m:
            return f"Number of flasks when seed average VCD is {m.group(1)} x10^6 cells/mL"
    if id1 == "BIO_ANTIFOAM":
        m = re.match(r"BOM_MAT(\d+)", u)
        if m:
            return f"Antifoam BOM material option {m.group(1)} (allowed part number)"
        m = re.match(r"(\d+)_MAX", id2)
        if m:
            return f"Maximum allowable antifoam quantity for part {m.group(1)}"
        m = re.match(r"(\d+)_MANUAL_REQ_QTY", id2)
        if m:
            return f"Manual required antifoam addition quantity for part {m.group(1)}"
    if id1 == "BIO_DAILY_SAMPLE":
        return DAILY.get(u, "")
    return ""

pval = []
catalog = {}
discarded = 0
for r in raw:
    client, product, id1, id2, typ, value, uom, sigfig = [clean(x) for x in r[:8]]
    keep = (client == "100") or (id1 == "BIO_DAILY_SAMPLE")
    if not keep:
        discarded += 1
        continue
    if id1 == "BIO_DAILY_SAMPLE":
        client, product = "100", "153"  # reclassify 903->100, MATNR PM_FG_100->153
    param_id = f"{id1}_{id2}"
    vt, tgt, lo, hi = val_shape(value)
    dt = data_type(typ, sigfig, vt, tgt, lo, hi)
    u = uop(id1)
    desc = describe(id1, id2, value)
    pval.append([param_id, u, "*", product, "3", vt, tgt, lo, hi, uom, "2026-01-01", "Released", sigfig])
    if param_id not in catalog:
        catalog[param_id] = [param_id, u, id1, role(id2), desc, dt, uom, vt,
                             "", "", "", "", "", "", "", (tgt if vt == "S" else f"{lo}-{hi}")]

wb = openpyxl.Workbook()
wb.remove(wb.active)
H = Font(bold=True, color="FFFFFF")
HF = PatternFill("solid", fgColor="2F4D2F")
NOTE = Font(bold=True)
EXTRA = PatternFill("solid", fgColor="FFF2CC")

def sheet(name, header, data, extra=0):
    s = wb.create_sheet(name)
    s.append(header)
    for c in s[1]:
        c.font = H
        c.fill = HF
    for row in data:
        s.append(row)
    if extra:
        for col in range(len(header) - extra + 1, len(header) + 1):
            s.cell(1, col).fill = EXTRA
    for i, col in enumerate(s.columns, 1):
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        s.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(w + 2, 9), 52)
    s.freeze_panes = "A2"
    return s

sheet("ZPP_PARAM",
      ["PARAM_ID", "UOP_ID", "PARAM_GROUP", "PARAM_ROLE", "PARAM_DESC", "DATA_TYPE", "UOM", "VALUE_TYPE",
       "IS_CALC", "FORMULA", "SRC_SYSTEM", "MAP_TYPE", "TOOL_READY", "IN_DTLS", "STEP_REF", "FORMAT_EXAMPLE"],
      list(catalog.values()))
sheet("ZPP_PVAL",
      ["PARAM_ID", "UOP_ID", "RESOURCE_ID", "MATNR", "SCOPE_LEVEL", "VAL_TYPE", "VAL_TARGET", "VAL_LOW",
       "VAL_HIGH", "UOM", "VALID_FROM", "STATUS", "SIG_FIG_SRC"],
      pval, extra=1)

ws = wb.create_sheet("Mapping & Assumptions")
rk = 1
def line(*vals, bold=False):
    global rk
    for j, v in enumerate(vals, 1):
        c = ws.cell(rk, j, v)
        if bold:
            c.font = NOTE
    rk += 1

line("Old 'Parameter Table Design'  ->  PPTable_Design_v2 format", bold=True); rk += 1
line("SCOPE (per user 2026-07-15): CLIENT 100 ONLY.", bold=True)
line("  Kept: client 100 / product 153 (FLASK_NUMBER + BIO_ANTIFOAM) + BIO_DAILY_SAMPLE.")
line("  BIO_DAILY_SAMPLE reclassified: client 903->100, MATNR PM_FG_100->153 (flagged - confirm MATNR).")
line(f"  Discarded {discarded} other client-903 rows (FG_XA_1 / PM_FG_100 non-daily-sample params)."); rk += 1
line("OLD column", "NEW target", "Notes", bold=True)
for a, b, c in [
    ("Product", "ZPP_PVAL.MATNR", "All 153 now (single client)."),
    ("Identifier1", "ZPP_PARAM.PARAM_GROUP", "Feature/XStep group (see crosswalk)."),
    ("Identifier2", "-> PARAM_ID", "PARAM_ID = Identifier1_Identifier2 (CHAR40)."),
    ("Type (N/C)", "DATA_TYPE", "C->STR; N & SIG_FIG=0->INT; N & >0->FLT."),
    ("Value", "VAL_TARGET (S) / VAL_LOW+VAL_HIGH (R)", "'a-b' auto-split into a Range."),
    ("UoM", "UOM", "Free text, unchanged."),
    ("SIG_FIG", "ZPP_PVAL.SIG_FIG_SRC (extension, yellow)", "No decimals col in v2; preserved."),
    ("(none)", "ZPP_PARAM.PARAM_DESC", "Authored as DRAFT from identifiers + XStep context - review."),
    ("(none)", "UOP_ID", "Old table had no unit-op axis. Inferred FLASK_NUMBER->INOCULUM, BIO_*->BIOREACTOR (placeholder)."),
]:
    line(a, b, c)
rk += 1
line("Kept as-is / blank:", bold=True)
line("  MIN_*/MAX_* kept as separate params (not merged into ranges).")
line("  Governance cols (SRC_SYSTEM/MAP_TYPE/TOOL_READY/IN_DTLS/STEP_REF) blank - absent in old table.")
line("  SCOPE_LEVEL=3, RESOURCE_ID='*' (values per-material; no resource/unit-op defaults in old data).")
line("  VALID_FROM=2026-01-01 / STATUS=Released are placeholders."); rk += 1
line("Identifier1 -> Phase 2 XStep", bold=True)
for k, v in GROUP_XSTEP.items():
    line(k, v, uop(k))
for i, col in enumerate(ws.columns, 1):
    w = max((len(str(c.value)) if c.value is not None else 0) for c in col)
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(w + 2, 10), 80)

sheet("Source (kept old rows)",
      ["Client", "Product", "Identifier1", "Identifier2", "Type", "Value", "UoM", "SIG_FIG"],
      [[clean(x) for x in r[:8]] for r in raw if (clean(r[0]) == "100" or clean(r[2]) == "BIO_DAILY_SAMPLE")])

wb.save(OUT)
print("wrote", OUT, "| catalog:", len(catalog), "| values:", len(pval), "| discarded:", discarded)
